import os
import string
import nltk
import openpyxl
import syllapy
import re


def stop_words():
    stop_words_all = set()
    for filename in os.listdir('./StopWords'):
        file_path = os.path.join('./StopWords', filename)
        with open(file_path, 'r', encoding='ISO-8859-1') as file:
            for line in file:
                stopword = line.strip().lower()
                stop_words_all.add(stopword)
    return sorted(stop_words_all)

def positive_words():
    positive_all = set()
    with open('./MasterDictionary/positive-words.txt', 'r', encoding='ISO-8859-1') as file:
        for line in file:
            word = line.strip().lower()
            positive_all.add(word)
    return sorted(positive_all)

def negative_words():
    negative_all = set()
    with open('./MasterDictionary/negative-words.txt', 'r', encoding='ISO-8859-1') as file:
        for line in file:
            word = line.strip().lower()
            negative_all.add(word)
    return sorted(negative_all)

def metrics_calculation(text):
    stopwords = stop_words()
    # print(len(stopwords))
    positivewords = positive_words()
    negativewords = negative_words()

    # Tokenize sentences and words
    sentences = nltk.sent_tokenize(text)
    words = nltk.word_tokenize(text)
    word_count = len(words)
    sentence_count = len(sentences)
    # print(words)
    # print(word_count)

    # Remove stopwords
    word_final = [word.lower() for word in words if word.lower() not in stopwords]
    word_final_count = len(word_final)
    # print(sentences)
    # print(word_final)
    # print(word_final_count)

    # Positive Score, Negative Score, Polarity Score, Subjectivity Score Calculation
    positive_score = len([word for word in word_final if word in positivewords])
    # print([word for word in word_final if word in positivewords])
    negative_score = len([word for word in word_final if word in negativewords])
    polarity_score = (positive_score - negative_score) / ((positive_score + negative_score) + 0.000006)
    subjective_score = (positive_score + negative_score) / (word_final_count + 0.000001)

    # Average Sentence Length
    avg_sentence_length = word_final_count / sentence_count

    # Average words per sentence
    avg_word_per_sentence = word_count / sentence_count

    # Word Count (Filtered)
    word_filtered = [word.lower().strip(string.punctuation) for word in words if word.lower().strip(string.punctuation) and word.lower() not in stopwords]
    words_filtered_count = len(word_filtered)
    # print(words_filtered_count)

    # Complex Words Percentage Calculation
    complex_word_count = len([word for word in word_final if syllapy.count(word) > 2])
    # print(len([word for word in word_final if syllapy.count(word) > 2]))
    percentage_complex = (complex_word_count / word_final_count) * 100.0

    # FOG index
    fog_index = 0.4 * (avg_sentence_length + percentage_complex)

    # Syllables per word
    syllable_per_word = sum(syllapy.count(word) for word in word_final) / word_final_count


     # Personal Pronouns Calculate
    pronouns = ['i', 'we', 'my', 'ours', 'us']
    text_joined = ' '.join(words)  
    pattern = r'\b(' + '|'.join(pronouns) + r')\b'
    personal_pronouns_count = len(re.findall(pattern, text_joined))
    # print(personal_pronouns_count)

    # Average Word Length
    avg_word_length = sum(len(word) for word in word_final) / word_final_count 


    return {
        "POSITIVE SCORE": positive_score,
        "NEGATIVE SCORE": negative_score,
        "POLARITY SCORE": polarity_score,
        "SUBJECTIVE SCORE": subjective_score,
        "AVG SENTENCE LENGTH": avg_sentence_length,
        "PERCENTAGE COMPLEX WORDS": percentage_complex,
        "FOG INDEX": fog_index,
        "AVG WORDS PER SENTENCE": avg_word_per_sentence,
        "COMPLEX WORD COUNT": complex_word_count,
        "FILTERED WORD COUNT": words_filtered_count,
        "SYLLABLE PER WORD": syllable_per_word,
        "PERSONAL PRONOUNS": personal_pronouns_count,
        "AVG WORD LENGTH": avg_word_length
    }
    

def process_articles():
    try:
        article_dir = '../Scrapping/articles'
        wb = openpyxl.load_workbook('../Output_Data.xlsx')
        sheet = wb.active
        row = 2
        for file in os.listdir(article_dir):
            file_path=os.path.join(article_dir,file)
            with open(file_path, 'r', encoding='ISO-8859-1') as f:
                text = f.read()
                ans = metrics_calculation(text)
                sheet.cell(row=row, column=3, value=ans["POSITIVE SCORE"])
                sheet.cell(row=row, column=4, value=ans["NEGATIVE SCORE"])
                sheet.cell(row=row, column=5, value=ans["POLARITY SCORE"])
                sheet.cell(row=row, column=6, value=ans["SUBJECTIVE SCORE"])
                sheet.cell(row=row, column=7, value=ans["AVG SENTENCE LENGTH"])
                sheet.cell(row=row, column=8, value=ans["PERCENTAGE COMPLEX WORDS"])
                sheet.cell(row=row, column=9, value=ans["FOG INDEX"])
                sheet.cell(row=row, column=10, value=ans["AVG WORDS PER SENTENCE"])
                sheet.cell(row=row, column=11, value=ans["COMPLEX WORD COUNT"])
                sheet.cell(row=row, column=12, value=ans["FILTERED WORD COUNT"])
                sheet.cell(row=row, column=13, value=ans["SYLLABLE PER WORD"])
                sheet.cell(row=row, column=14, value=ans["PERSONAL PRONOUNS"])
                sheet.cell(row=row, column=15, value=ans["AVG WORD LENGTH"])
                print(f'Created Sheet at row {row}')

                row+=1

        wb.save('../Output_Data_updated.xlsx')
    except Exception as e:
        print(Exception)
process_articles()
